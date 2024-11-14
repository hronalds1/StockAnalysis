import streamlit as st
import requests
from bs4 import BeautifulSoup
import time
import random
import os
import json
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

TICKER_FILE = "last_tickers.txt"
LISTS_FILE = "saved_lists.json"


def load_tickers():
    if os.path.exists(TICKER_FILE):
        with open(TICKER_FILE, "r") as f:
            content = f.read().strip()
            return [ticker.strip() for ticker in content.split(",") if ticker.strip()]
    return []


def save_tickers(tickers):
    with open(TICKER_FILE, "w") as f:
        f.write(",".join(tickers))


def load_lists():
    if os.path.exists(LISTS_FILE):
        with open(LISTS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_lists(lists):
    with open(LISTS_FILE, "w") as f:
        json.dump(lists, f)


def get_stock_ratios(tickers):
    results = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
    }

    for ticker in tickers:
        try:
            analysis_url = f"https://finance.yahoo.com/quote/{ticker}/analysis?p={ticker}"
            stats_url = f"https://finance.yahoo.com/quote/{ticker}/key-statistics?p={ticker}"

            response_analysis = requests.get(analysis_url, headers=headers)
            response_stats = requests.get(stats_url, headers=headers)

            soup_analysis = BeautifulSoup(response_analysis.text, 'html.parser')
            soup_stats = BeautifulSoup(response_stats.text, 'html.parser')

            def extract_value(soup, label, alt_labels=None):
                try:
                    # Debug print for PEG ratio search
                    if 'PEG' in label:
                        print(f"\nSearching for PEG ratio in HTML...")
                        for table in soup.find_all('table'):
                            for row in table.find_all('tr'):
                                print(f"Row text: {row.get_text()}")

                    # Method 1: Direct table cell search with exact match
                    for table in soup.find_all('table'):
                        for row in table.find_all('tr'):
                            cells = row.find_all(['td', 'th'])
                            for i, cell in enumerate(cells):
                                cell_text = cell.get_text().strip().lower()
                                if label.lower() in cell_text:  # Changed to 'in' instead of exact match
                                    if i + 1 < len(cells):
                                        value = cells[i + 1].get_text().strip()
                                        if 'PEG' in label:
                                            print(f"Found PEG value: {value}")
                                        return value

                    # Method 2: Try alternative labels if provided
                    if alt_labels:
                        for alt_label in alt_labels:
                            for table in soup.find_all('table'):
                                for row in table.find_all('tr'):
                                    cells = row.find_all(['td', 'th'])
                                    for i, cell in enumerate(cells):
                                        if alt_label.lower() in cell.get_text().strip().lower():
                                            if i + 1 < len(cells):
                                                return cells[i + 1].get_text().strip()

                    return 'N/A'
                except Exception as e:
                    print(f"Error extracting {label} for {ticker}: {str(e)}")
                    return 'N/A'

            # Extract all values with debug output
            print(f"\nExtracting data for {ticker}...")

            forward_pe = extract_value(soup_stats, 'Forward P/E')
            print(f"Forward P/E: {forward_pe}")

            # Try multiple variations for PEG ratio
            yahoo_peg = extract_value(soup_stats, 'PEG Ratio (5 yr expected)')
            if yahoo_peg == 'N/A':
                yahoo_peg = extract_value(soup_stats, 'PEG Ratio')
            print(f"Yahoo PEG: {yahoo_peg}")

            price_to_sales = extract_value(soup_stats, 'Price/Sales')
            qtrly_rev_growth_yoy = extract_value(soup_stats, 'Quarterly Revenue Growth')
            one_year_growth = extract_value(soup_analysis, 'Next Year', ['Growth Estimate Next Year'])
            five_year_growth = extract_value(soup_analysis, 'Next 5 Years (per annum)',
                                             ['Growth Est Next 5Y', '5 Year Growth Est', 'Next Five Years'])

            # Calculate our own PEG ratio
            def calculate_peg(pe, growth):
                try:
                    if pe != 'N/A' and growth != 'N/A':
                        pe_val = float(pe.replace(',', ''))
                        growth_val = float(growth.replace('%', '').replace(',', ''))
                        if growth_val != 0:
                            return f"{(pe_val / growth_val):.2f}"
                except:
                    pass
                return 'N/A'

            calculated_peg = calculate_peg(forward_pe, five_year_growth)

            # Calculate QoQ revenue growth using yfinance
            stock = yf.Ticker(ticker)
            financials = stock.quarterly_financials

            if not financials.empty and 'Total Revenue' in financials.index:
                revenue = financials.loc['Total Revenue']
                if len(revenue) >= 2:
                    latest_revenue = revenue.iloc[0]
                    previous_revenue = revenue.iloc[1]
                    qoq_growth = ((latest_revenue - previous_revenue) / previous_revenue) * 100
                    qoq_growth_str = f"{qoq_growth:.2f}%"
                else:
                    qoq_growth_str = "N/A"
            else:
                qoq_growth_str = "N/A"

            def format_value(value):
                if value == "N/A":
                    return ""
                try:
                    # Handle percentage values
                    if isinstance(value, str) and '%' in value:
                        num = float(value.replace('%', '').replace(',', '').strip())
                        return f"{num:.2f}%"
                    # Handle numeric values
                    num = float(value.replace(',', '').strip())
                    return f"{num:.2f}"
                except (ValueError, AttributeError):
                    return value

            # Print detailed metrics for this stock
            print(f"\nDetailed metrics for {ticker}:")
            print(f"Forward P/E: {forward_pe}")
            print(f"5-Year Growth Rate: {five_year_growth}")
            print(f"Yahoo PEG Ratio: {yahoo_peg}")
            print(f"Our Calculated PEG: {calculated_peg}")
            if yahoo_peg != 'N/A' and calculated_peg != 'N/A':
                try:
                    diff = abs(float(yahoo_peg) - float(calculated_peg))
                    print(f"PEG Difference: {diff:.2f}")
                except:
                    pass

            stock_data = [
                ticker,
                format_value(forward_pe),
                format_value(yahoo_peg),
                format_value(calculated_peg),
                format_value(price_to_sales),
                format_value(qtrly_rev_growth_yoy),
                qoq_growth_str if qoq_growth_str != "N/A" else "",
                format_value(one_year_growth),
                format_value(five_year_growth)
            ]

        except Exception as e:
            print(f"Error fetching data for {ticker}: {str(e)}")
            stock_data = [ticker, "", "", "", "", "", "", "", ""]

        results.append(stock_data)
        time.sleep(random.uniform(2, 4))

    return results
def edit_tickers(tickers):
    print("\nCurrent tickers:", " ".join(tickers))

    remove_input = input(
        "Remove Symbols by typing 'all' or individual symbols (if any) separated by spaces (press Enter to skip): ").strip().upper()
    if remove_input == 'ALL':
        tickers = []
    elif remove_input:
        to_remove = remove_input.split()
        tickers = [ticker for ticker in tickers if ticker not in to_remove]

    add_input = input(
        "Add Symbols by listing symbols (if any) separated by spaces (press Enter to skip): ").strip().upper()
    if add_input:
        new_tickers = add_input.split()
        tickers.extend([ticker for ticker in new_tickers if ticker not in tickers])

    return tickers

def save_current_list(tickers, lists):
    list_name = input("Enter a name for this list: ").strip()
    if list_name:
        lists[list_name] = tickers
        save_lists(lists)
        print(f"List '{list_name}' saved successfully.")
    else:
        print("List not saved. No name provided.")

def choose_list(lists):
    if not lists:
        print("No saved lists found.")
        return None

    print("\nSaved lists:")
    for i, (name, tickers) in enumerate(lists.items(), 1):
        print(f"{i}. {name}: {', '.join(tickers)}")

    choice = input("\nEnter the number of the list you want to use (or press Enter to skip): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(lists):
        chosen_list = list(lists.keys())[int(choice) - 1]
        return lists[chosen_list]
    return None

def create_new_list():
    new_tickers = input("Enter tickers for the new list (space-separated): ").strip().upper()
    return [ticker.strip() for ticker in new_tickers.split() if ticker.strip()]

def save_list_prompt(tickers, lists):
    save_choice = input("Do you want to save this list? (y/n): ").strip().lower()
    if save_choice == 'y':
        save_current_list(tickers, lists)

def delete_list(lists):
    if not lists:
        print("No saved lists found.")
        return lists

    print("\nSaved lists:")
    for i, name in enumerate(lists.keys(), 1):
        print(f"{i}. {name}")

    choice = input("\nEnter the number of the list you want to delete (or press Enter to cancel): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(lists):
        list_to_delete = list(lists.keys())[int(choice) - 1]
        del lists[list_to_delete]
        save_lists(lists)
        print(f"List '{list_to_delete}' has been deleted.")
    return lists

def save_to_file(stock_ratios, headers):
    filename = input("Enter the filename to save the data (e.g., stock_analysis.xlsx): ").strip()
    if not filename:
        filename = "stock_analysis.xlsx"
    elif not filename.endswith('.xlsx'):
        filename += '.xlsx'

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"

        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add data
        for row, stock_data in enumerate(stock_ratios, start=2):
            for col, value in enumerate(stock_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=len(stock_ratios) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        wb.save(filename)
        print(f"\nData has been saved to {filename}")

        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            print(f"File '{filename}' has been successfully created and contains data.")
        else:
            print(f"Warning: File '{filename}' may be empty or not created properly.")

    except Exception as e:
        print(f"An error occurred while saving the file: {str(e)}")

def main():
    tickers = load_tickers()
    lists = load_lists()

    while True:
        if not tickers:
            print("\nNo stocks in the current list. You must create a list of stocks to look up.")
            while not tickers:
                tickers = create_new_list()
                if not tickers:
                    print("The list is still empty. You must enter at least one stock symbol.")
            save_tickers(tickers)
            save_list_prompt(tickers, lists)

        print(f"\nCurrent list is: {' '.join(tickers)}")
        print("1. Create new list")
        print("2. Choose saved list")
        print("3. Edit list")
        print("4. Delete list")
        print("5. Look up ratios")
        print("6. Quit")

        choice = input("Enter your choice (1-6): ").strip()

        if choice == '1':
            new_tickers = create_new_list()
            if new_tickers:
                tickers = new_tickers
                save_tickers(tickers)
                save_list_prompt(tickers, lists)
            else:
                print("List creation cancelled. Current list remains unchanged.")
        elif choice == '2':
            chosen_tickers = choose_list(lists)
            if chosen_tickers:
                tickers = chosen_tickers
                save_tickers(tickers)
                print(f"Current list updated: {' '.join(tickers)}")
            else:
                tickers = []
        elif choice == '3':
            tickers = edit_tickers(tickers)
            save_tickers(tickers)
        elif choice == '4':
            lists = delete_list(lists)
            if not lists and not tickers:
                print("All lists deleted. You need to create a new list.")
                tickers = []
        elif choice == '5':
            if not tickers:
                print("The ticker list is empty. Please create a new list.")
                continue
            print(f"\nFetching data for: {' '.join(tickers)}")
            print("Please wait...\n")
            stock_ratios = get_stock_ratios(tickers)
            headers = ["Ticker", "Forward P/E", "Yahoo PEG", "Calc PEG", "Price/Sales",
                      "Qtrly Rev Growth YoY", "Qtrly Rev Growth QoQ", "1Y Rev Growth Est",
                      "5Y Rev Growth Est"]

            # Print the data in a tabular format
            print("\n{:<10} {:<12} {:<10} {:<10} {:<12} {:<20} {:<20} {:<20} {:<20}".format(*headers))
            print("-" * 134)
            for row in stock_ratios:
                print("{:<10} {:<12} {:<10} {:<10} {:<12} {:<20} {:<20} {:<20} {:<20}".format(*row))

            print("\nNote: Data is fetched from Yahoo Finance.")
            print("If you see blank fields, it means the data wasn't available or there was an issue fetching it.")

            save_option = input("\nDo you want to save this data to an Excel file? (y/n): ").strip().lower()
            if save_option == 'y':
                save_to_file(stock_ratios, headers)
        elif choice == '6':
            print("Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
